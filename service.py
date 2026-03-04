from __future__ import annotations

from datetime import datetime, date, time, timedelta
from typing import Optional, Dict, List, Tuple, Set
import re
from io import BytesIO

from sqlalchemy import text
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from .schemas import (
    AdminAttendanceSettings,
    SummaryEmployeeMetric,
    AttendanceDetailsResponse,
    DailyDetailRow,
)

SETTINGS_KEY = "admin_attendance_settings"


def _get_table_columns(db: Session, table_name: str, schema: str = "public") -> Set[str]:
    rows = db.execute(
        text(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = :schema AND table_name = :table
            """
        ),
        {"schema": schema, "table": table_name},
    ).mappings().all()
    return {r["column_name"] for r in rows}


def _get_settings_row(db: Session) -> Optional[dict]:
    try:
        row = db.execute(
            text("SELECT key, value_json FROM settings WHERE key = :k"),
            {"k": SETTINGS_KEY},
        ).mappings().first()
        return dict(row) if row else None
    except Exception:
        return None


def get_settings(db: Session) -> AdminAttendanceSettings:
    row = _get_settings_row(db)
    if not row:
        return AdminAttendanceSettings()
    try:
        import json

        payload = json.loads(row["value_json"] or "{}")
        return AdminAttendanceSettings(**payload)
    except Exception:
        return AdminAttendanceSettings()


def save_settings(db: Session, settings: AdminAttendanceSettings) -> AdminAttendanceSettings:
    import json

    payload = json.dumps(settings.model_dump(mode="json"), ensure_ascii=False)
    db.execute(
        text(
            """
            INSERT INTO settings(key, value_json)
            VALUES(:k, :v)
            ON CONFLICT (key) DO UPDATE SET value_json = EXCLUDED.value_json
            """
        ),
        {"k": SETTINGS_KEY, "v": payload},
    )
    db.commit()
    return settings


def _apply_tz_like(src: datetime, target: datetime) -> datetime:
    if src.tzinfo is not None and target.tzinfo is None:
        return target.replace(tzinfo=src.tzinfo)
    if src.tzinfo is None and target.tzinfo is not None:
        return target.replace(tzinfo=None)
    return target


def _minutes_between(a: datetime, b: datetime) -> int:
    b = _apply_tz_like(a, b)
    if b < a:
        return 0
    return int((b - a).total_seconds() // 60)


def _assume_is_night(settings: AdminAttendanceSettings, start_at: datetime) -> bool:
    h = start_at.hour
    return h >= settings.assume_night_start_hour_from or h <= settings.assume_night_start_hour_to


def _auto_checkout_end_at(
    settings: AdminAttendanceSettings, start_at: datetime, work_date: date, shift_type: Optional[str]
) -> datetime:
    # shift_type이 NIGHT면 우선
    if shift_type and str(shift_type).upper() in {s.upper() for s in settings.night_shift_types}:
        return _apply_tz_like(
            start_at, datetime.combine(work_date + timedelta(days=1), settings.night_regular_checkout_time)
        )

    is_night = _assume_is_night(settings, start_at)
    if not is_night:
        return _apply_tz_like(start_at, datetime.combine(work_date, settings.day_regular_checkout_time))

    return _apply_tz_like(start_at, datetime.combine(work_date + timedelta(days=1), settings.night_regular_checkout_time))


def _fetch_users(db: Session) -> List[dict]:
    cols = _get_table_columns(db, "users")
    # 내부직원(관리자/운영자/회사직원)만 집계 대상
    # role_id: 6(관리자), 7(운영자), 8(회사직원)
    role_filter_sql = ""
    if "role_id" in cols:
        role_filter_sql = " AND role_id IN (6, 7, 8) "

    if "is_active" in cols:
        return db.execute(
            text(
                "SELECT id as user_id, name as user_name "
                "FROM users "
                "WHERE is_active = true" + role_filter_sql + "ORDER BY id ASC"
            )
        ).mappings().all()

    return db.execute(
        text(
            "SELECT id as user_id, name as user_name "
            "FROM users "
            "WHERE 1=1" + role_filter_sql + "ORDER BY id ASC"
        )
    ).mappings().all()



def _select_sessions_sql(settings: AdminAttendanceSettings, db: Session) -> str:
    ws_cols = _get_table_columns(db, "work_sessions")

    def sel(col: str, cast: str, alias: str) -> str:
        return f"ws.{col} as {alias}" if col in ws_cols else f"NULL::{cast} as {alias}"

    return f"""
        SELECT
            ws.user_id as user_id,
            u.name as user_name,
            ws.start_at as start_at,
            ws.end_at as end_at,
            ws.work_date_basis as work_date_basis,
            {sel(settings.session_type_column, "text", "session_type")},
            {sel(settings.shift_type_column, "text", "shift_type")},
            {sel(settings.place_column, "text", "place")},
            {sel(settings.task_column, "text", "task")},
            {sel(settings.is_holiday_column, "boolean", "is_holiday")}
        FROM work_sessions ws
        JOIN users u ON u.id = ws.user_id
    """


def fetch_summary_report(
    db: Session, start_date: date, end_date: date, user_id: Optional[int]
) -> Tuple[List[SummaryEmployeeMetric], AdminAttendanceSettings]:
    settings = get_settings(db)
    users = _fetch_users(db)
    if user_id is not None:
        users = [u for u in users if int(u["user_id"]) == int(user_id)]

    sql = _select_sessions_sql(settings, db) + """
        WHERE ws.start_at >= :start_dt
          AND ws.start_at < :end_dt
          AND (:uid IS NULL OR ws.user_id = :uid)
        ORDER BY ws.user_id ASC, ws.start_at ASC
    """

    sessions = db.execute(
        text(sql),
        {
            "start_dt": datetime.combine(start_date, time(0, 0)),
            "end_dt": datetime.combine(end_date + timedelta(days=1), time(0, 0)),
            "uid": user_id,
        },
    ).mappings().all()

    # 연차 총량(있으면)
    ent_map: Dict[int, int] = {}
    try:
        ents = db.execute(text("SELECT user_id, annual_leave_total FROM leave_entitlements")).mappings().all()
        ent_map = {int(e["user_id"]): int(e["annual_leave_total"]) for e in ents}
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass

    office_set = {s.upper() for s in settings.office_session_types}
    offsite_set = {s.upper() for s in settings.offsite_session_types}
    leave_set = {s.upper() for s in settings.leave_session_types}
    half_leave_set = {s.upper() for s in settings.half_leave_session_types}

    agg: Dict[int, dict] = {
        int(u["user_id"]): {
            "user_id": int(u["user_id"]),
            "user_name": u["user_name"],
            "total_minutes": 0,
            "days_with_work": set(),
            "overtime_days": set(),
            "holiday_work_days": set(),
            "annual_leave_used": 0,
            "annual_leave_used_this_period": 0,
            "half_leave_used": 0,
            "offsite_count": 0,
            "office_count": 0,
            "offsite_places": set(),

            # ✅ 추가 업무(일) 산정을 위한 일자별 근무 분 합계
            "daily_work_minutes": {},  # work_date -> minutes
            "extra_work_days": set(),
        }
        for u in users
    }

    for s in sessions:
        uid = int(s["user_id"])
        if uid not in agg:
            continue

        start_at: datetime = s["start_at"]
        end_at: Optional[datetime] = s["end_at"]
        # ✅ 날짜 집계 기준은 work_date_basis 우선(야간/자정넘김 케이스 안정화)
        work_date = s.get("work_date_basis") or start_at.date()

        session_type = str(s.get("session_type") or "")
        shift_type = s.get("shift_type")
        place = str(s.get("place") or "")
        is_holiday = bool(s.get("is_holiday")) if s.get("is_holiday") is not None else False

        if end_at is None:
            end_at = _auto_checkout_end_at(settings, start_at, work_date, shift_type)
        else:
            end_at = _apply_tz_like(start_at, end_at)

        # ✅ 월차/반차 세션은 근무시간 합산에서 제외(시간이 들어가면 총근무시간이 왜곡될 수 있음)
        st_u = session_type.upper()
        is_leave = st_u in leave_set or st_u in half_leave_set

        minutes = _minutes_between(start_at, end_at)
        if not is_leave:
            agg[uid]["total_minutes"] += minutes
            agg[uid]["days_with_work"].add(work_date)

            # ✅ 일자별 근무 분 누적(세션 합산 기반)
            dmap = agg[uid]["daily_work_minutes"]
            dmap[work_date] = int(dmap.get(work_date, 0)) + int(minutes)

        if not is_leave:
            # 휴일 근무는 세션 단위가 아니라 '일자' 단위로만 집계
            if is_holiday:
                agg[uid]["holiday_work_days"].add(work_date)
        if st_u in leave_set:
            agg[uid]["annual_leave_used"] += 1
            agg[uid]["annual_leave_used_this_period"] += 1
        if st_u in half_leave_set:
            agg[uid]["half_leave_used"] += 1
        if st_u in offsite_set:
            agg[uid]["offsite_count"] += 1
            if place:
                agg[uid]["offsite_places"].add(place)
        if st_u in office_set:
            agg[uid]["office_count"] += 1

    # ✅ 야근/추가업무(일) 계산 규칙(서로 배타)
    # - 야근: 야근인정기준 초과(>) AND 추가업무인정기준 이하(<=)
    # - 추가업무: 추가업무인정기준 초과(>)
    overtime_thr_min = int(float(settings.overtime_threshold_hours or 0) * 60)
    extra_thr_min = int(float(settings.extra_work_threshold_hours or 0) * 60)

    for uid, a in agg.items():
        dmap: Dict[date, int] = a.get("daily_work_minutes", {})
        for d, mins in dmap.items():
            m = int(mins)
            if extra_thr_min > 0 and m > extra_thr_min:
                a["extra_work_days"].add(d)
                # 추가업무인 날은 야근으로 카운트하지 않음
                continue
            if overtime_thr_min > 0 and m > overtime_thr_min:
                a["overtime_days"].add(d)

    items: List[SummaryEmployeeMetric] = []
    for uid, a in agg.items():
        total_minutes = int(a["total_minutes"])
        total_hours = round(total_minutes / 60.0, 2) if total_minutes else 0.0
        days = len(a["days_with_work"]) or 0
        avg = round(total_hours / days, 2) if days else 0.0
        annual_total = int(ent_map.get(uid, 15))
        annual_used = int(a["annual_leave_used"])

        items.append(
            SummaryEmployeeMetric(
                user_id=uid,
                user_name=a["user_name"],
                total_work_minutes=total_minutes,
                total_work_hours=total_hours,
                avg_work_hours=avg,
                overtime_days=len(a["overtime_days"]),
                holiday_work_days=len(a["holiday_work_days"]),
                extra_work_days=len(a["extra_work_days"]),
                annual_leave_total=annual_total,
                annual_leave_used=annual_used,
                annual_leave_used_this_period=int(a["annual_leave_used_this_period"]),
                half_leave_used=int(a["half_leave_used"]),
                offsite_count=int(a["offsite_count"]),
                office_count=int(a["office_count"]),
                offsite_places=sorted(list(a["offsite_places"])),
            )
        )

    return items, settings


def fetch_details(db: Session, start_date: date, end_date: date, user_id: int) -> AttendanceDetailsResponse:
    settings = get_settings(db)

    user_row = db.execute(
        text("SELECT id as user_id, name as user_name FROM users WHERE id = :uid"),
        {"uid": user_id},
    ).mappings().first()
    if not user_row:
        raise ValueError("user not found")

    sql = _select_sessions_sql(settings, db) + """
        WHERE ws.start_at >= :start_dt
          AND ws.start_at < :end_dt
          AND ws.user_id = :uid
        ORDER BY ws.start_at ASC
    """

    sessions = db.execute(
        text(sql),
        {
            "start_dt": datetime.combine(start_date, time(0, 0)),
            "end_dt": datetime.combine(end_date + timedelta(days=1), time(0, 0)),
            "uid": user_id,
        },
    ).mappings().all()

    by_day: Dict[date, DailyDetailRow] = {}

    office_set = {s.upper() for s in settings.office_session_types}
    offsite_set = {s.upper() for s in settings.offsite_session_types}
    leave_set = {s.upper() for s in settings.leave_session_types}
    half_leave_set = {s.upper() for s in settings.half_leave_session_types}

    for s in sessions:
        start_at: datetime = s["start_at"]
        end_at: Optional[datetime] = s["end_at"]

        # ✅ 날짜 집계 기준은 work_date_basis 우선 (야간/자정넘김 케이스를 안정적으로 처리)
        work_date = s.get("work_date_basis") or start_at.date()

        shift_type = s.get("shift_type")
        session_type = str(s.get("session_type") or "")
        place = str(s.get("place") or "")
        task = str(s.get("task") or "")
        is_holiday = bool(s.get("is_holiday")) if s.get("is_holiday") is not None else False

        if end_at is None:
            end_at = _auto_checkout_end_at(settings, start_at, work_date, shift_type)
        else:
            end_at = _apply_tz_like(start_at, end_at)

        st_u = str(session_type or "").upper()
        is_leave = st_u in leave_set or st_u in half_leave_set

        minutes = _minutes_between(start_at, end_at)

        row = by_day.get(work_date)
        if not row:
            row = DailyDetailRow(work_date=work_date)
            by_day[work_date] = row

        # ✅ 월차/반차 세션은 근무시간/출퇴근(첫/마지막) 산정에서 제외
        if not is_leave:
            if row.first_start_at is None or start_at < row.first_start_at:
                row.first_start_at = start_at
            if row.last_end_at is None or end_at > row.last_end_at:
                row.last_end_at = end_at

            row.work_minutes += minutes
            row.work_hours = round(row.work_minutes / 60.0, 2)

        if shift_type is not None and row.shift_type is None:
            row.shift_type = str(shift_type)
        if is_holiday:
            row.is_holiday = True

        if session_type and session_type not in row.session_types:
            row.session_types.append(session_type)
        if place and place not in row.places:
            row.places.append(place)
        if task and task not in row.tasks:
            row.tasks.append(task)

        # ✅ 월차/반차 표시용(해당 유형 세션의 최초 start_at)
        if st_u in leave_set:
            if row.leave_start_at is None or start_at < row.leave_start_at:
                row.leave_start_at = start_at
        if st_u in half_leave_set:
            if row.half_leave_start_at is None or start_at < row.half_leave_start_at:
                row.half_leave_start_at = start_at

    days = sorted(by_day.values(), key=lambda r: r.work_date)

    # ✅ 추가 업무 인정 플래그
    extra_thr_min = int(float(settings.extra_work_threshold_hours or 0) * 60)
    if extra_thr_min > 0:
        for r in days:
            r.extra_work_recognized = int(r.work_minutes or 0) > extra_thr_min

    return AttendanceDetailsResponse(
        user_id=int(user_row["user_id"]),
        user_name=str(user_row["user_name"]),
        start_date=start_date,
        end_date=end_date,
        days=days,
    )


def build_excel(db: Session, start_date: date, end_date: date, user_id: Optional[int]) -> bytes:
    """엑셀 생성: 요약 + 상세 2시트"""
    items, settings = fetch_summary_report(db, start_date=start_date, end_date=end_date, user_id=user_id)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "요약"
    ws1.append(["직원","총 근무시간","평균 근무시간(일)","야근수(일)","휴일 근무(일)","추가 업무 인정(일)","월차 사용(기간)","월차 사용(누적)","총 월차","반차","외근","사무실","외근 장소"])
    for it in items:
        ws1.append([it.user_name,it.total_work_hours,it.avg_work_hours,it.overtime_days,it.holiday_work_days,it.extra_work_days,
                    it.annual_leave_used_this_period,it.annual_leave_used,it.annual_leave_total,it.half_leave_used,
                    it.offsite_count,it.office_count,", ".join(it.offsite_places or [])])
    _style_sheet(ws1)

    ws2 = wb.create_sheet("상세")
    ws2.append(["직원","날짜","주야","휴일","근무시간","추가업무(Y)","근무형태(session_type)","외근 장소","업무/사유(task)"])

    sql = _select_sessions_sql(settings, db) + """
        WHERE ws.start_at >= :start_dt
          AND ws.start_at < :end_dt
          AND (:uid IS NULL OR ws.user_id = :uid)
        ORDER BY ws.user_id ASC, ws.start_at ASC
    """
    sessions = db.execute(
        text(sql),
        {
            "start_dt": datetime.combine(start_date, time(0, 0)),
            "end_dt": datetime.combine(end_date + timedelta(days=1), time(0, 0)),
            "uid": user_id,
        },
    ).mappings().all()

    grouped: Dict[Tuple[int, date], dict] = {}
    for s in sessions:
        uid = int(s["user_id"])
        uname = str(s["user_name"])
        start_at: datetime = s["start_at"]
        end_at: Optional[datetime] = s["end_at"]
        # ✅ 날짜 집계 기준은 work_date_basis 우선(야간/자정넘김 안정화)
        d = s.get("work_date_basis") or start_at.date()
        shift_type = s.get("shift_type")
        session_type = str(s.get("session_type") or "")
        place = str(s.get("place") or "")
        task = str(s.get("task") or "")
        is_holiday = bool(s.get("is_holiday")) if s.get("is_holiday") is not None else False

        if end_at is None:
            end_at = _auto_checkout_end_at(settings, start_at, d, shift_type)
        else:
            end_at = _apply_tz_like(start_at, end_at)

        # ✅ 월차/반차 세션은 근무시간/추가업무 산정에서 제외(리포트/화면과 동일 기준)
        st_u = str(session_type or "").upper()
        leave_set = {x.upper() for x in settings.leave_session_types}
        half_leave_set = {x.upper() for x in settings.half_leave_session_types}
        is_leave = st_u in leave_set or st_u in half_leave_set

        minutes = _minutes_between(start_at, end_at)

        key = (uid, d)
        row = grouped.get(key)
        if not row:
            row = {"user_name": uname, "date": d, "shift_type": str(shift_type) if shift_type else "",
                   "is_holiday": is_holiday, "minutes": 0, "session_types": set(), "places": set(), "tasks": []}
            grouped[key] = row

        if not is_leave:
            row["minutes"] += minutes
        if is_holiday:
            row["is_holiday"] = True
        if shift_type and not row["shift_type"]:
            row["shift_type"] = str(shift_type)
        if session_type:
            row["session_types"].add(session_type)
        if place:
            row["places"].add(place)
        if task:
            row["tasks"].append(task)

    extra_thr_min = int(float(getattr(settings, "extra_work_threshold_hours", 0) or 0) * 60)

    for (_uid, _d), row in sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1])):
        hours = round(row["minutes"] / 60.0, 2)
        is_extra = "Y" if extra_thr_min > 0 and int(row["minutes"]) > extra_thr_min else ""
        ws2.append([
            row["user_name"],
            row["date"].isoformat(),
            row["shift_type"],
            "Y" if row["is_holiday"] else "",
            hours,
            is_extra,
            ", ".join(sorted(list(row["session_types"]))),
            ", ".join(sorted(list(row["places"]))),
            " / ".join(row["tasks"])[:3000],
        ])

    _style_sheet(ws2, wrap_columns={7, 8})

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _style_sheet(ws, wrap_columns: Set[int] = set()):
    bold = Font(bold=True)
    center = Alignment(vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 16
        if col in wrap_columns:
            ws.column_dimensions[letter].width = 44

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = wrap if c in wrap_columns else Alignment(vertical="center")


# =========================
# 관리자 근태 정정(출근/퇴근 날짜+시간)
# =========================

def get_day_sessions(db: Session, user_id: int, work_date: date) -> dict:
    """work_date_basis 기준으로 해당 일자의 원장(work_sessions) 목록을 반환"""
    ws_cols = _get_table_columns(db, "work_sessions")
    sel_cols = ["id"]
    for c in ["session_type", "shift_type", "start_at", "end_at", "work_date_basis", "place", "task", "is_holiday", "source"]:
        if c in ws_cols:
            sel_cols.append(c)

    rows = db.execute(
        text(
            f"""
            SELECT {", ".join(sel_cols)}
            FROM work_sessions
            WHERE user_id = :uid AND work_date_basis = :d
            ORDER BY start_at ASC NULLS LAST, id ASC
            """
        ),
        {"uid": user_id, "d": work_date},
    ).mappings().all()

    sessions = [dict(r) for r in rows]

    # ✅ 프론트 정정(수정) 모달 프리필: 첫 출근 / 마지막 퇴근 (날짜+시간)
    settings = get_settings(db)
    first_start_at: Optional[datetime] = None
    last_end_at: Optional[datetime] = None

    for s in sessions:
        sa = s.get("start_at")
        ea = s.get("end_at")
        st = s.get("shift_type")
        if not sa:
            continue

        if first_start_at is None or sa < first_start_at:
            first_start_at = sa

        # end_at이 없으면 자동퇴근 인정(end) 계산
        if ea is None:
            eff_end = _auto_checkout_end_at(settings, sa, work_date, st)
        else:
            eff_end = _apply_tz_like(sa, ea)

        s["effective_end_at"] = eff_end
        s["work_minutes"] = _minutes_between(sa, eff_end)

        if last_end_at is None or eff_end > last_end_at:
            last_end_at = eff_end

    return {
        "user_id": user_id,
        "work_date": work_date,
        "sessions": sessions,
        "first_start_at": first_start_at,
        "last_end_at": last_end_at,
    }


def _parse_hm(hm: str) -> time:
    hm = (hm or "").strip()
    if not re.match(r"^\d{2}:\d{2}$", hm):
        raise ValueError("시간 형식은 HH:MM 입니다.")
    h, m = hm.split(":")
    return time(int(h), int(m))


def _combine_kst(d: date, hm: str) -> datetime:
    # DB는 timestamptz이며, 기존 데이터도 +09:00 기반 저장이 많아서 KST로 생성
    tz = datetime.now().astimezone().tzinfo
    t = _parse_hm(hm)
    return datetime(d.year, d.month, d.day, t.hour, t.minute, 0, tzinfo=tz)


def _log_correction_best_effort(
    db: Session,
    user_id: int,
    work_date: date,
    action: str,
    before_json: str,
    after_json: str,
    reason: str,
    editor_user_id: int,
) -> None:
    """attendance_corrections 스키마가 프로젝트마다 달라서, 컬럼 존재 여부를 보고 가능한 만큼만 기록."""
    try:
        cols = _get_table_columns(db, "attendance_corrections")
    except Exception:
        return

    payload = {
        "user_id": user_id,
        "target_date": work_date,
        "work_date": work_date,
        "action": action,
        "before_json": before_json,
        "after_json": after_json,
        "reason": reason,
        "editor_user_id": editor_user_id,
    }

    # 요청 변경사항 json (있으면)
    if "requested_changes_json" in cols:
        try:
            import json as _json
            payload["requested_changes_json"] = _json.dumps(
                {
                    "action": action,
                    "target_date": work_date.isoformat(),
                    "reason": reason,
                    "before": before_json,
                    "after": after_json,
                },
                ensure_ascii=False,
            )
        except Exception:
            payload["requested_changes_json"] = after_json

    # status 컬럼(예전 스키마) 있으면 기본값
    if "status" in cols and "status" not in payload:
        payload["status"] = "PENDING"

    # created_at 등은 default가 있으면 생략

    # 실제 insert할 컬럼/파라미터만 골라서
    ins_cols = [c for c in payload.keys() if c in cols]
    if not ins_cols:
        return

    params = {c: payload[c] for c in ins_cols}

    db.execute(
        text(
            f"""
            INSERT INTO attendance_corrections({", ".join(ins_cols)})
            VALUES({", ".join([":" + c for c in ins_cols])})
            """
        ),
        params,
    )


def apply_day_correction(
    db: Session,
    *,
    user_id: int,
    work_date: date,
    start_date: Optional[date],
    start_hm: Optional[str],
    end_date: Optional[date],
    end_hm: Optional[str],
    reason: str,
    editor_user_id: int,
) -> None:
    """
    - work_sessions에서 work_date_basis=work_date인 세션을 찾는다.
    - 첫 세션(start_at 최소) => 출근(start_at) 정정 대상
    - 마지막 세션(end_at 최대 또는 start_at 최대) => 퇴근(end_at) 정정 대상
    """
    sessions = db.execute(
        text(
            """
            SELECT id, start_at, end_at
            FROM work_sessions
            WHERE user_id=:uid AND work_date_basis=:d
            ORDER BY start_at ASC NULLS LAST, id ASC
            """
        ),
        {"uid": user_id, "d": work_date},
    ).mappings().all()

    if not sessions:
        raise ValueError("해당 날짜에 수정할 세션이 없습니다. (work_date_basis 기준)")

    before = [dict(r) for r in db.execute(
        text(
            """
            SELECT id, session_type, shift_type, start_at, end_at, work_date_basis, place, task, is_holiday, source
            FROM work_sessions
            WHERE user_id=:uid AND work_date_basis=:d
            ORDER BY start_at ASC NULLS LAST, id ASC
            """
        ),
        {"uid": user_id, "d": work_date},
    ).mappings().all()]

    first_id = sessions[0]["id"]

    # last: choose row with greatest COALESCE(end_at, start_at)
    last_row = max(sessions, key=lambda r: (r["end_at"] or r["start_at"] or datetime.min))
    last_id = last_row["id"]

    if start_date and start_hm:
        new_start = _combine_kst(start_date, start_hm)
        db.execute(text("UPDATE work_sessions SET start_at=:v WHERE id=:id"), {"v": new_start, "id": first_id})

    if end_date and end_hm:
        new_end = _combine_kst(end_date, end_hm)
        db.execute(text("UPDATE work_sessions SET end_at=:v WHERE id=:id"), {"v": new_end, "id": last_id})

    db.commit()

    after = [dict(r) for r in db.execute(
        text(
            """
            SELECT id, session_type, shift_type, start_at, end_at, work_date_basis, place, task, is_holiday, source
            FROM work_sessions
            WHERE user_id=:uid AND work_date_basis=:d
            ORDER BY start_at ASC NULLS LAST, id ASC
            """
        ),
        {"uid": user_id, "d": work_date},
    ).mappings().all()]

    try:
        import json as _json
        before_json = _json.dumps(before, default=str, ensure_ascii=False)
        after_json = _json.dumps(after, default=str, ensure_ascii=False)
    except Exception:
        before_json = str(before)
        after_json = str(after)

    # best-effort logging (실패해도 정정은 유지)
    try:
        action = "SET_START_END_DATETIME"
        _log_correction_best_effort(db, user_id, work_date, action, before_json, after_json, reason, editor_user_id)
        db.commit()
    except Exception:
        db.rollback()
        # ignore
        return
